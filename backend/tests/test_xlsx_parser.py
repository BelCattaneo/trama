import io
from pathlib import Path

import pytest
from openpyxl import Workbook

from trama.parsing.xlsx_parser import ParseError, parse_xlsx

FIXTURES = Path(__file__).parent / "fixtures" / "parsing"


def _read(name: str) -> io.BytesIO:
    return io.BytesIO((FIXTURES / name).read_bytes())


def test_standard_five_rows_no_warnings():
    payload = parse_xlsx(_read("standard.xlsx"))
    assert len(payload.lines) == 5
    assert payload.warnings == []
    products = [line.product for line in payload.lines]
    assert products == [
        "Zanahoria",
        "Tomate perita",
        "Lechuga",
        "Manzana",
        "Pera",
    ]
    quantities = [line.quantity for line in payload.lines]
    assert quantities == [3.0, 5.0, 2.0, 4.0, 6.0]
    units = [line.unit for line in payload.lines]
    assert units == ["kg", "kg", "atado", "kg", "kg"]


def test_standard_raw_text_includes_full_row():
    payload = parse_xlsx(_read("standard.xlsx"))
    assert payload.lines[0].raw_text == "Zanahoria | 3 | kg"


def test_alternate_headers_in_row_three():
    payload = parse_xlsx(_read("alternate_headers.xlsx"))
    assert len(payload.lines) == 5
    assert payload.warnings == []
    products = [line.product for line in payload.lines]
    assert products == ["Acelga", "Tomate cherry", "Naranja", "Limón", "Frutilla"]
    # Tomate cherry should have its full label preserved (variety preservation)
    assert "Tomate cherry" in products


def test_malformed_skips_invalid_and_emits_warnings():
    payload = parse_xlsx(_read("malformed.xlsx"))
    products = [line.product for line in payload.lines]
    # only Zanahoria (row 2) and Manzana (row 6, comma decimal) survive
    assert products == ["Zanahoria", "Manzana"]
    assert any("sin cantidad" in w for w in payload.warnings)
    assert any("cantidad inválida" in w for w in payload.warnings)
    assert any("sin producto" in w for w in payload.warnings)


def test_malformed_comma_decimal_parsed_as_float():
    payload = parse_xlsx(_read("malformed.xlsx"))
    manzana = next(line for line in payload.lines if line.product == "Manzana")
    assert manzana.quantity == 1.5


def test_malformed_warnings_include_row_numbers():
    payload = parse_xlsx(_read("malformed.xlsx"))
    # standard row 1 = header; row 2 = Zanahoria OK; row 3 = empty qty;
    # row 4 = invalid qty; row 5 = blank (skipped silently);
    # row 6 = Manzana OK; row 7 = empty product
    joined = " ".join(payload.warnings)
    assert "fila 3" in joined
    assert "fila 4" in joined
    assert "fila 7" in joined


def test_malformed_empty_row_skipped_silently():
    payload = parse_xlsx(_read("malformed.xlsx"))
    # Empty row at index 5 must not produce a warning
    assert not any("fila 5" in w for w in payload.warnings)


def test_no_recognizable_headers_raises():
    with pytest.raises(ParseError) as exc:
        parse_xlsx(_read("no_headers.xlsx"))
    assert "no se encontraron columnas reconocidas" in str(exc.value)


def test_corrupt_zip_raises_parse_error():
    # Magic prefix of xlsx but body is not a valid zip
    corrupt = io.BytesIO(b"PK\x03\x04" + b"xl/" + b"\x00" * 20)
    with pytest.raises(ParseError) as exc:
        parse_xlsx(corrupt)
    assert "archivo xlsx inválido" in str(exc.value)


def test_quantity_parsing_handles_int_float_string():
    # Build a tiny xlsx on the fly

    wb = Workbook()
    ws = wb.active
    ws.append(["Producto", "Cantidad", "Unidad"])
    ws.append(["a", 3, "kg"])       # int
    ws.append(["b", 1.5, "kg"])     # float
    ws.append(["c", "1.5", "kg"])   # string with dot
    ws.append(["d", "  2  ", "kg"]) # string with whitespace
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    payload = parse_xlsx(buf)
    quantities = [line.quantity for line in payload.lines]
    assert quantities == [3.0, 1.5, 1.5, 2.0]


def test_unit_optional_when_missing_column():

    wb = Workbook()
    ws = wb.active
    ws.append(["Producto", "Cantidad"])
    ws.append(["Zanahoria", 3])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    payload = parse_xlsx(buf)
    assert len(payload.lines) == 1
    assert payload.lines[0].unit is None


def test_unit_optional_when_cell_empty():

    wb = Workbook()
    ws = wb.active
    ws.append(["Producto", "Cantidad", "Unidad"])
    ws.append(["Zanahoria", 3, ""])
    ws.append(["Tomate", 5, None])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    payload = parse_xlsx(buf)
    assert all(line.unit is None for line in payload.lines)
