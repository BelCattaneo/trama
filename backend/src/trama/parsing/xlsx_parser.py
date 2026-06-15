from typing import BinaryIO

from openpyxl import load_workbook

from trama.parsing.columns import REQUIRED_FIELDS, canonicalize_columns
from trama.parsing.schema import ParseLine, ParsePayload

_HEADER_SCAN_ROWS = 5


class ParseError(Exception):
    """Raised when the parser cannot extract any usable lines."""


def _parse_quantity(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None


def _row_to_strings(row) -> list[str]:
    return ["" if c is None else str(c) for c in row]


def _row_is_empty(row) -> bool:
    return all(c is None or str(c).strip() == "" for c in row)


def _find_header(rows: list[list]) -> tuple[int, dict[str, str]]:
    """Scan the first N rows for one whose headers cover REQUIRED_FIELDS.
    Returns (header_row_index, {raw_header: canonical_field}).
    """
    for idx, row in enumerate(rows[:_HEADER_SCAN_ROWS]):
        if _row_is_empty(row):
            continue
        mapping = canonicalize_columns(_row_to_strings(row))
        if REQUIRED_FIELDS.issubset(set(mapping.values())):
            return idx, mapping
    raise ParseError("no se encontraron columnas reconocidas")


def parse_xlsx(stream: BinaryIO) -> ParsePayload:
    """Parse an .xlsx file stream into a ParsePayload.
    Raises ParseError if no recognizable header row is found.
    """
    workbook = load_workbook(stream, read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        raise ParseError("no se encontraron columnas reconocidas")

    rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    if not rows:
        raise ParseError("no se encontraron columnas reconocidas")

    header_idx, mapping = _find_header(rows)
    header_strings = _row_to_strings(rows[header_idx])
    canonical_by_col_idx: dict[int, str] = {}
    for col_idx, header in enumerate(header_strings):
        if header in mapping:
            canonical_by_col_idx[col_idx] = mapping[header]

    lines: list[ParseLine] = []
    warnings: list[str] = []

    for row_offset, row in enumerate(rows[header_idx + 1 :], start=1):
        row_num = header_idx + 1 + row_offset  # 1-indexed for users
        if _row_is_empty(row):
            continue

        fields: dict[str, object] = {}
        for col_idx, canonical in canonical_by_col_idx.items():
            if col_idx < len(row):
                fields[canonical] = row[col_idx]

        product = fields.get("product")
        product_str = "" if product is None else str(product).strip()
        if not product_str:
            warnings.append(f"fila {row_num} sin producto, saltada")
            continue

        raw_qty = fields.get("quantity")
        if raw_qty is None or (isinstance(raw_qty, str) and not raw_qty.strip()):
            warnings.append(f"fila {row_num} sin cantidad, saltada")
            continue
        quantity = _parse_quantity(raw_qty)
        if quantity is None:
            warnings.append(
                f"fila {row_num} con cantidad inválida: '{raw_qty}', saltada"
            )
            continue

        unit_value = fields.get("unit")
        unit = (
            str(unit_value).strip()
            if unit_value is not None and str(unit_value).strip()
            else None
        )

        raw_text = " | ".join(_row_to_strings(row)).strip(" |")

        lines.append(
            ParseLine(
                product=product_str,
                quantity=quantity,
                unit=unit,
                raw_text=raw_text or None,
            )
        )

    return ParsePayload(lines=lines, warnings=warnings)
